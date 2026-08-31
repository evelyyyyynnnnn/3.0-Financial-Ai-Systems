#!/usr/bin/env python3
"""Generate project-overview.xlsx — every project in the repository.

    python scripts/build_overview.py

One table (ROWS) is the source of truth: site, project, core idea, what makes
it non-obvious, status, stack, entry point, and source-file count. Edit that
table when a project is added, retired, or changes status, then re-run.

The Summary sheet counts by status with COUNTIF/SUMIF over the Projects sheet
rather than hardcoded totals, so it stays correct when ROWS changes. openpyxl
writes formulas without cached values, so tools that read cached values see
blanks until the file is opened in Excel (which recalculates on open) or run
through a LibreOffice recalculation.
"""
import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
INK   = Font(name=FONT, size=10)
BOLD  = Font(name=FONT, size=10, bold=True)
HEAD  = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=FONT, size=14, bold=True, color="1F3864")
SUB   = Font(name=FONT, size=9, italic=True, color="595959")
MONO  = Font(name="Consolas", size=9)

HFILL = PatternFill("solid", fgColor="1F3864")
BAND  = PatternFill("solid", fgColor="F2F5F9")
thin  = Side(style="thin", color="D6DCE4")
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP   = Alignment(vertical="top", wrap_text=True)
TOPL  = Alignment(vertical="top", wrap_text=True, horizontal="left")
CTR   = Alignment(vertical="center", horizontal="center", wrap_text=True)

STATUS_FILL = {
    "Built":       PatternFill("solid", fgColor="E2EFDA"),
    "Built (live)":PatternFill("solid", fgColor="C6E0B4"),
    "Research":    PatternFill("solid", fgColor="DDEBF7"),
    "Retired":     PatternFill("solid", fgColor="EDEDED"),
    "Third-party": PatternFill("solid", fgColor="FCE4D6"),
    "Planned":     PatternFill("solid", fgColor="FFF2CC"),
}

# ---- Every project in the repository. Facts verified against the tree. ----
# (Site, Project, Core idea, What makes it non-obvious, Status, Stack, Entry, Files)
ROWS = [
 ("project-1", "Portfolio Optimization Engine",
  "Learn a portfolio allocation policy with reinforcement learning instead of solving a static optimisation once.",
  "Three agents share one environment — a PPO written from scratch in TensorFlow, plus Stable-Baselines3 SAC (continuous weights) and DQN (discretised grid) — so the same problem is attacked by three different learning regimes and compared.",
  "Built", "Python, TensorFlow, Stable-Baselines3, Gymnasium, Plotly", "src/main.py", 9),

 ("project-1", "Financial Network Risk",
  "Treat financial risk as a graph problem: build a knowledge graph of entities and relationships, then train GNNs over it.",
  "Propagation paths are structural, not pairwise. Correlation between two firms says nothing about the third firm that connects them, which is what a graph representation recovers.",
  "Built", "Python, PyTorch, GNN, NetworkX", "main.py", 6),

 ("project-1", "Volatility Forecasting",
  "Forecast market volatility with LSTM and Transformer models over engineered technical indicators.",
  "Volatility is the uncertainty term every downstream risk model consumes, so the pipeline is built around calibration: early stopping, Optuna hyperparameter search, and persisted metrics rather than a single fitted model.",
  "Built", "Python, PyTorch, Transformer, LSTM, Optuna, Docker", "LSTM-Volatility-Prediction/main.py", 14),

 ("project-1", "Credit Risk AI",
  "Model credit risk from three modalities at once — text, market data, and images — rather than from financial ratios alone.",
  "Credit deterioration usually shows up in language before it shows up in numbers. Fusing a text branch with market and visual signals is an attempt to catch the earlier signal.",
  "Built", "Python, PyTorch, Transformers, OpenCV, scikit-learn", "train.py", 5),

 ("project-1", "Eventized Microstructure LLM",
  "Tokenise order-book events into sequences so transformer architectures can model market microstructure as a language problem.",
  "Reframes microstructure from a time-series problem into a sequence-modelling one, which lets slippage, volatility spikes and liquidity crises be predicted with the machinery built for text.",
  "Research", "LLM, Transformer, Market Microstructure", "Frame-work/Logic-Framework.md", 1),

 ("project-1", "Live Trading Engine",
  "Execute factor strategies against the Alpaca API under real latency constraints, in parallel C++ and Python implementations.",
  "The execution layer that makes the rest real: a strategy is only a decision framework once it survives microsecond-level order management, risk limits, and live WebSocket market data.",
  "Built", "C++17, CMake, Boost, WebSocket++, Python, Alpaca API", "C++/main.cpp", 38),

 ("project-1", "Trading Simulation Platform",
  "One platform for market data, portfolio state, strategy backtesting, and investment commentary, with real-time updates.",
  "Backtesting is the verification step: a strategy claim is only trustworthy once it survives out-of-sample. Absorbed the retired Investment Analysis Dashboard, so it is now the single trading front end.",
  "Built", "Node.js, Express, Socket.IO, Chart.js, Jest", "server.js", 20),

 ("project-1", "Giant Portfolio Tracker",
  "Make institutional 13F positioning legible and current — searchable by fund, manager, and fund type.",
  "Market transparency in the literal sense. A scheduled job pulls two Notion databases monthly and commits data.json, which redeploys the site; the page itself never changes.",
  "Built (live)", "Static HTML, Notion API, GitHub Actions, Python, Vercel", "giant-portfolio/index.html", 1),

 ("project-1", "Macroeconomic Data Pipeline",
  "A notebook that assembles and cleans macroeconomic series for downstream models.",
  "The shared input layer — the macro variables several of the other projects condition on.",
  "Research", "Jupyter, Python, pandas", "code-pipeline/macroeconomic-data-pipeline.ipynb", 1),

 ("project-2", "Filing Intelligence",
  "Report what CHANGED in a company's SEC risk disclosures since its prior filing, rather than summarising the filing.",
  "Scores by CONTAINMENT, not similarity. difflib's ratio() is symmetric, so a risk factor kept word-for-word and then extended with a new admission scores low purely because the new sentence is longer — and the escalation gets misfiled as a brand-new disclosure. Containment asks how much of the prior sentence survives inside the current one, so the extended version is correctly read as the same disclosure, escalated.",
  "Built", "Python (stdlib only), SEC EDGAR API, static JS", "pipeline/build_dataset.py", 6),

 ("project-2", "Contagion Observatory",
  "Measure which crypto-equity links actually TRANSMIT stress, and simulate what a shock to one asset does to the rest.",
  "Correlation is the wrong instrument: two assets can move together every day and transmit nothing — they share a factor. Edge strength weights TAIL DEPENDENCE at 0.65 — whether a pair fails together — which surfaces channels that ordinary correlation ranks as unremarkable.",
  "Built", "Python (stdlib only), Stooq, CoinGecko, SVG", "pipeline/build_dataset.py", 5),

 ("project-2", "Contract Audit",
  "Detect known Solidity vulnerability classes and report each one so a human auditor can confirm or dismiss it.",
  "Every finding states the conditions under which THAT DETECTOR IS WRONG. A tool that hides its failure modes costs more review time than it saves. Validated both ways: six planted bugs all caught at the right lines, and a safe contract with the same surface reports zero false positives.",
  "Built", "Python (stdlib only), Solidity static analysis", "analyzer/build_report.py", 5),

 ("roadmap", "Private Credit Data Provenance",
  "Extract terms from private credit documents where every extracted value carries a citation back to the source span.",
  "For a figure feeding a risk model or a regulatory filing, uncitable is unusable. LLM extraction produces a clean table and destroys the audit trail doing it.",
  "Planned", "— (scoping document only)", "roadmap/private-credit-data-provenance/README.md", 0),

 ("roadmap", "Tokenized Fixed-Income Analytics",
  "Measure liquidity, holder concentration, and redemption latency for tokenized debt from on-chain trade data.",
  "Conventional fixed-income analytics run against a new feed. Distinguishes realised secondary depth from wash volume, and observed redemption time from the prospectus.",
  "Planned", "— (scoping document only)", "roadmap/tokenized-fixed-income-analytics/README.md", 0),

 ("project-1/archive", "Investment Analysis Dashboard",
  "Investment dashboard with live market data and AI-generated commentary. Retired after the portfolio audit.",
  "The Trading Simulation Platform covered it — 26 endpoints against 4, with market summary and sector performance duplicated. Its two unique pieces were migrated rather than lost: advisory generation and four research pages.",
  "Retired", "Node.js, Express, Chart.js, Tailwind", "archive/trading-system-dashboard/", 0),

 ("project-1/reference", "Options Volatility Trading",
  "Long/short straddle volatility strategy against Interactive Brokers, with an alerting bot and Flask dashboard.",
  "NOT ORIGINAL WORK. MIT-licensed, Copyright (c) 2021 MCF Long Short — a group project from Union University's Masters in Computational Finance. Its ib_client/ is IBKR's official TWS API. Kept in reference/ so the distinction is structural, not a caveat.",
  "Third-party", "Python, Flask, IBKR TWS API", "reference/options-volatility-trading/", 0),
]

HEADERS = ["Site", "Project", "Core idea", "What makes it non-obvious",
           "Status", "Stack", "Entry point", "Source files"]
WIDTHS  = [17, 30, 52, 62, 13, 34, 34, 11]

wb = openpyxl.Workbook()

# ============ Sheet 1: Projects ============
ws = wb.active
ws.title = "Projects"
ws["A1"] = "3.0 Financial AI Systems — Project Overview"
ws["A1"].font = TITLE
ws.merge_cells("A1:H1")
ws["A2"] = ("Every project in the repository, including retired, third-party, and planned work. "
            "Facts verified against the source tree; 'Source files' counts .py/.js/.cpp/.h/.sol, "
            "excluding node_modules and vendored libraries.")
ws["A2"].font = SUB
ws.merge_cells("A2:H2")
ws.row_dimensions[2].height = 26

r = 4
for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = HEAD, HFILL, CTR, BOX
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[r].height = 22

for i, row in enumerate(ROWS):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=rr, column=c, value=val)
        cell.border = BOX
        cell.alignment = CTR if c in (1, 5, 8) else TOPL
        cell.font = MONO if c == 7 else INK
        if c == 5:
            cell.fill = STATUS_FILL.get(val, BAND)
            cell.font = BOLD
        elif i % 2 == 1:
            cell.fill = BAND
    ws.row_dimensions[rr].height = 92

last = r + len(ROWS)
ws.auto_filter.ref = f"A{r}:H{last}"
ws.freeze_panes = f"C{r+1}"

# ============ Sheet 2: Summary (formulas, not hardcoded) ============
s2 = wb.create_sheet("Summary")
s2["A1"] = "Counts by status"
s2["A1"].font = TITLE
s2["A2"] = "Every figure below is a COUNTIF over the Projects sheet, so it updates if that sheet changes."
s2["A2"].font = SUB
s2.merge_cells("A2:D2")

s2["A4"], s2["B4"], s2["C4"] = "Status", "Projects", "Source files"
for c in "ABC":
    cell = s2[f"{c}4"]
    cell.font, cell.fill, cell.alignment, cell.border = HEAD, HFILL, CTR, BOX
s2.column_dimensions["A"].width = 18
s2.column_dimensions["B"].width = 12
s2.column_dimensions["C"].width = 14
s2.column_dimensions["D"].width = 58

order = ["Built (live)", "Built", "Research", "Planned", "Retired", "Third-party"]
NOTE = {
 "Built (live)": "Deployed and refreshing on a schedule.",
 "Built":        "Runnable code with a documented entry point.",
 "Research":     "Paper or notebook; no runnable service.",
 "Planned":      "Scoping document only — no implementation.",
 "Retired":      "Superseded; kept for reference, not counted as current work.",
 "Third-party":  "Not original work. Retained and labelled as reference.",
}
for i, st in enumerate(order):
    rr = 5 + i
    s2.cell(row=rr, column=1, value=st).font = BOLD
    s2.cell(row=rr, column=1).fill = STATUS_FILL[st]
    s2.cell(row=rr, column=2, value=f'=COUNTIF(Projects!$E$5:$E${last},A{rr})')
    s2.cell(row=rr, column=3, value=f'=SUMIF(Projects!$E$5:$E${last},A{rr},Projects!$H$5:$H${last})')
    s2.cell(row=rr, column=4, value=NOTE[st]).font = INK
    for c in range(1, 5):
        s2.cell(row=rr, column=c).border = BOX
        if c in (2, 3):
            s2.cell(row=rr, column=c).font = INK
            s2.cell(row=rr, column=c).alignment = CTR

tr = 5 + len(order)
s2.cell(row=tr, column=1, value="Total").font = BOLD
s2.cell(row=tr, column=2, value=f"=SUM(B5:B{tr-1})").font = BOLD
s2.cell(row=tr, column=3, value=f"=SUM(C5:C{tr-1})").font = BOLD
for c in range(1, 4):
    s2.cell(row=tr, column=c).border = BOX
    s2.cell(row=tr, column=c).alignment = CTR

s2[f"A{tr+2}"] = "Deployable sites"
s2[f"A{tr+2}"].font = BOLD
for i, (name, root, pages) in enumerate([
    ("Portfolio", "project-1/", "Overview + 7 project pages + /giant-portfolio/"),
    ("Trustworthy Systems", "project-2/", "Landing + /filing-intelligence/ + /contagion/ + /contract-audit/"),
]):
    rr = tr + 3 + i
    s2.cell(row=rr, column=1, value=name).font = INK
    s2.cell(row=rr, column=2, value=root).font = MONO
    s2.cell(row=rr, column=4, value=pages).font = INK
    for c in range(1, 5):
        s2.cell(row=rr, column=c).border = BOX

s2[f"A{tr+6}"] = ("Note: sample data currently ships with Filing Intelligence and Contagion so both "
                  "render before a live pull — fictional issuers in one, simulated series in the other, "
                  "labelled on the page. Contract Audit analyses real Solidity from disk.")
s2[f"A{tr+6}"].font = SUB
s2.merge_cells(f"A{tr+6}:D{tr+7}")
s2[f"A{tr+6}"].alignment = TOPL

OUT = pathlib.Path(__file__).resolve().parent.parent / "project-overview.xlsx"
wb.save(OUT)
print(f"Wrote {OUT} — {len(ROWS)} projects")
